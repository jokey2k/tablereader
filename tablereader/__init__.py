#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
    tablereader
    ~~~~~~~~~~~

    Unified abstraction for handling xls, xlsx and CSV files in Python, reads tables as unicode

    :copyright: 12.2015 by Markus Ullmann, mail@markus-ullmann.de
"""

#
# Python imports
from datetime import datetime
import csv

#
# environment imports
import openpyxl
import openpyxl.cell.cell
from six import next as six_next, PY2, string_types as six_string_types
import xlrd2

#
# local imports
import tablereader._csv_from_pypy as _csv

#
# constants
CLEAR_STRING = ""  # used to speed up processing in pypy, has no functional meaning

__version__ = "1.1.1"


class BaseXLReader(object):
    """Mimic _csv.reader interface for DictReader to be able to handle an xls sheet"""

    def __init__(self, filename, sheetname=None):
        self.line_num = 0
        self.filename = filename
        self.sheetname = sheetname
        self._reader = xlrd2.open_workbook(filename)
        self.sheetnames = self._reader.sheet_names()
        if sheetname is None:
            self._sheet = self._reader.sheet_by_index(0)
        else:
            if sheetname not in self.sheetnames:
                raise ValueError("No such sheet %s" % sheetname)
            self._sheet = self._reader.sheet_by_name(sheetname)

    def __iter__(self):
        return self

    def __next__(self):
        return self.next()

    def next(self):
        try:
            items = self.stringified_row()
        except IndexError:
            raise StopIteration
        self.line_num += 1
        return items


class XLReaderPy2(BaseXLReader):

    def stringified_row(self):
        """Ensure row contents are all strings"""

        row = self._sheet.row_values(self.line_num)
        newrow = []
        for element in row:
            if isinstance(element, float):
                element = str(element)
            newrow.append(unicode(element) if not isinstance(element, unicode) else element)
        return newrow


class XLReaderPy3(BaseXLReader):

    def stringified_row(self):
        """Ensure row contents are all strings"""

        row = self._sheet.row_values(self.line_num)
        newrow = []
        for element in row:
            if isinstance(element, float):
                element = str(element)
            newrow.append(element)
        return newrow


if PY2:
    XLReader = XLReaderPy2
else:
    XLReader = XLReaderPy3


class XLSXReader(object):
    """Mimic _csv.reader interface for DictReader to be able to handle an xlsx sheet"""

    def __init__(self, filename, sheetname=None):
        self.line_num = 0
        self.filename = filename
        self.sheetname = sheetname
        self._reader = openpyxl.load_workbook(filename, read_only=True)
        self.sheetnames = self._reader.sheetnames
        if sheetname is None:
            self._sheet = self._reader[self.sheetnames[0]]
        else:
            self._sheet = self._reader.get_sheet_by_name(sheetname)
            if self._sheet is None:
                raise ValueError("No such sheet %s" % sheetname)
        self._iter = self._sheet.iter_rows()

    def __iter__(self):
        return self

    def __next__(self):
        return self.next()

    def next(self):
        items = self.stringified_row()
        self.line_num += 1
        return items

    def stringified_row(self):
        """Ensure row contents are all strings"""

        row = six_next(self._iter)
        newrow = []
        for element in row:
            if element.value is not None:
                if isinstance(element.value, datetime):
                    element = str(element.value)
                elif element.data_type is openpyxl.cell.cell.TYPE_NUMERIC:
                    element = str(element.value)
                else:
                    element = element.value
            else:
                element = CLEAR_STRING
            newrow.append(element)
        return newrow


class CSVStrippingReader(object):
    """Patches CSV reader to allow for whitespace stripping"""

    def __init__(self, reader):
        self.reader = reader

    def __getattr__(self, name):
        if name == "reader":
            return self.reader
        else:
            return getattr(self.reader, name)

    def __iter__(self):
        return self

    def __next__(self):
        return self.next()

    def next(self):
        row = self.reader.next()
        return [entry.strip() for entry in row]


class TableReader(object):
    """Consolidated interface for reading csv, xls and xlsx"""

    def __init__(self, filename, sheet=None, fieldnames=None, strip_whitespaces=False, force_type=None, delimiter=";", quotechar='"'):
        """Optional parameter description:

        :sheet: Sheet to use if document has sheet support, otherwise defaults to first one
        :fieldnames: If the table has no columns, column names in fixed order may be given here. Falls back to first row otherwise
        :strip_whitespaces: before a column is passed on, a strip() on the string is be performed
        :force_type: Enforce a certain file format, if filename's extension is not matching. Options are "CSV", "XLS" and "XLSX"
        :delimiter: Delimiter used on CSV reading

        """

        # Sanitize input
        if force_type:
            force_type = force_type.lower().strip()

        self.strip_whitespaces = strip_whitespaces
        self.manually_strip_whitespaces = strip_whitespaces
        self.is_stringio = "StringI" in filename.__class__.__name__
        if self.is_stringio and force_type is None:
                raise ValueError("StringIO given but no forced type, I cannot guess!")
        if self.is_stringio:
            self.filehandle = filename
        else:
            self.filehandle = open(filename)
        self.reader = csv.DictReader(self.filehandle, delimiter=delimiter, quotechar=quotechar, fieldnames=fieldnames)
        if force_type == "csv" or (force_type is None and filename.endswith(".csv")):
            # if it is plain csv, no action to take for us
            if strip_whitespaces:
                self.reader.reader = CSVStrippingReader(self.reader.reader)
                self.manually_strip_whitespaces = False
        elif force_type == "unicodecsv":
            self.reader.reader = _csv.reader(self.filehandle, delimiter=delimiter, quotechar=quotechar)
            if strip_whitespaces:
                self.reader.reader = CSVStrippingReader(self.reader.reader)
                self.manually_strip_whitespaces = False
        elif force_type == "xls" or (force_type is None and filename.endswith(".xls")):
            # Monkey patch reader to use XLS mimic sheet reader instead
            self.reader.reader = XLReader(filename, sheet)
        elif force_type == "xlsx" or (force_type is None and (filename.endswith(".xlsx") or
                                                              filename.endswith(".xlsm") or
                                                              filename.endswith(".xltx") or
                                                              filename.endswith(".xltm"))):
            # Monkey patch reader to use XLSX mimic sheet reader instead
            self.reader.reader = XLSXReader(filename, sheet)
        else:
            raise NotImplementedError("Unsupported file extension and no known type given as parameter")

    def __iter__(self):
        return self

    def __next__(self):
        return self.next()

    def __del__(self):
        if not self.is_stringio:
            self.filehandle.close()

    def next(self):
        # Strip whitespace here if the reader was not able to do it by itself already (only CSV is capable of doing it currently)
        if self.manually_strip_whitespaces:
            newrow = {}
            for k, v in six_next(self.reader).iteritems():
                if isinstance(v, six_string_types):
                    newrow[k] = v.strip()
                else:
                    newrow[k] = v
            return newrow
        else:
            return six_next(self.reader)

    @property
    def line_num(self):
        return self.reader.line_num

    @property
    def fieldnames(self):
        return self.reader.fieldnames

    @staticmethod
    def get_sheet_names(filename):
        if filename.endswith(".xls"):
            reader = xlrd2.open_workbook(filename)
            return reader.sheet_names()
        elif (filename.endswith(".xlsx") or
              filename.endswith(".xlsm") or
              filename.endswith(".xltx") or
              filename.endswith(".xltm")):
            reader = openpyxl.load_workbook(filename, read_only=True)
            return reader.get_sheet_names()
        else:
            raise NotImplementedError("Unsupported file format")

    def close(self):
        if not self.is_stringio:
            self.filehandle.close()


class OffsetTableReader(TableReader):
    """Tablereader able to treat a row with a special value as header row. Just give search string as second parameter"""

    def __init__(self, filename, header_start_content, sheet=None, fieldnames=[], strip_whitespaces=False, force_type=None):
        super(OffsetTableReader, self).__init__(filename=filename, sheet=sheet, fieldnames=fieldnames, strip_whitespaces=strip_whitespaces, force_type=force_type)

        def iter_columns():
            for row in self.reader.reader:
                for columnno in range(len(row)):
                    yield row, columnno

        if not fieldnames:
            for row, columnno in iter_columns():
                check_column = row[columnno]
                if check_column == header_start_content:
                    self.reader.fieldnames = [columnname for columnname in row]
                    break
